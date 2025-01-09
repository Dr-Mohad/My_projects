import pandas as pd
import datetime
import time
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
def register_students():
    # Load the data
    main_data = pd.read_excel('Registeration.xlsx', dtype={'Student_Number': str, 'Guardian_Number': str})
    qw = pd.read_excel('Teacher registeration.xlsx' , dtype={'Number': str})
    Teacher_Number = qw.Number.tolist() 
    i_s = main_data.ID.tolist()
    n = main_data.Student_Number.tolist()
    n2 = main_data.Guardian_Number.tolist()
    continue_registration = True

    # Function to generate a new unique ID
    def generate_next_id(existing_ids):
        if not existing_ids:
            return 'MS001'
        max_id = max(int(i[2:]) for i in existing_ids if i.startswith('MS'))
        return f"MS{max_id + 1:03}"

    while continue_registration:
        try:
            # Ask how many students to register
            num_students = int(input('How many students would you like to register? '))
        except ValueError:
            print('Please enter a valid number.')
            continue

        # Loop through the number of students to register
        for _ in range(num_students):
            # Automatically generate a new ID
            iD = generate_next_id(i_s)

            print(f"The new ID for the student is: {iD}")

            # Validate Full Name
            while True:
                full_name = input('Enter Full Name (First Middle Last): ').title().strip().split()
                if all(len(i) >= 3  and i.isalpha() for i in full_name) and len(full_name) == 3:
                    break
                else:
                    print("The full name is invalid. Ensure it is three words, each at least 4 characters long.")

            # Validate Student Number
            while True:
                student_number = input('Enter Student Number: ')
                if student_number in n or student_number in n2 or student_number in Teacher_Number:
                    print('This number is already registered. Please try again.')
                elif student_number.startswith('061') and len(student_number) == 10 and student_number.isdigit():
                    break
                else:
                    print("Invalid student number. Please ensure it starts with '061' and is 10 digits long.")

            # Validate Gender
            while True:
                gender = input('Enter Gender (Male/Female): ').capitalize()
                if gender in ['Male', 'Female']:
                    break
                else:
                    print("Please enter either Male or Female.")

            # Validate Guardian Name
            while True:
                guardian_name = input('Enter Guardian Name (Full Name): ').title().strip().split()
                if guardian_name == full_name:
                    print('The student and the guardian cannot be the same. Please try again.')
                elif all(len(part) >= 4  and part.isalpha() for part in guardian_name) and len(guardian_name) == 3 :
                    break
                else:
                    print("Please enter a valid guardian full name with three words.")

            # Validate Guardian Number
            while True:
                guardian_number = input('Enter Guardian Number: ')
                if guardian_number in n: 
                    print('This number is already registered. Please try again.')
                elif guardian_number == student_number:
                    print('This number is already exist. Please try again.')
                elif guardian_number.startswith('061') and len(guardian_number) == 10 and guardian_number.isdigit() :
                    break
                else:
                    print("Invalid guardian number. Please ensure it starts with '061' and is 10 digits long.")

            # Validate Date of Birth
            while True:
                dob = input('Enter Date of Birth (dd/mm/yyyy): ')
                try:
                    day, month, year = map(int, dob.split('/'))
                    date = f'{day}/{month}/{year}'
                    current_year = datetime.datetime.now().year
                    if datetime.datetime.strptime(str(date),'%d/%m/%Y') and year <=(current_year-7):
                        break
                    else:
                        print("Invalid date of birth. Please try again.")
                except ValueError:
                    print("Please use the correct format: dd/mm/yyyy.")

            # Validate Class
            while True:
                print('1. Form One\n2. Form Two\n3. Form Three\n4. Form Four')
                class_choice = input('Enter Class: ')
                if class_choice == '1':
                    class_name = 'Form One'
                    sheet_name = 'Form1'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '2':
                    class_name = 'Form Two'
                    sheet_name = 'Form2'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '3':
                    class_name = 'Form Three'
                    sheet_name = 'Form3'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                elif class_choice == '4':
                    class_name = 'Form Four'
                    sheet_name = 'Form4'
                    attendance_data = pd.read_excel('Attendence.xlsx', sheet_name=sheet_name)
                    exam_data = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                    break
                else:
                    print('Please enter a valid number.')

            # Create a new student record
            new_student = {
                'ID': iD,
                'Full_Name': " ".join(full_name),
                'Student_Number': student_number,
                'Gender': gender,
                'Guardian_Name': " ".join(guardian_name),
                'Guardian_Number': guardian_number,
                'Date_of_Birth': dob,
                'Registration_Date': time.strftime('%d/%m/%Y'),
                'Class': class_name
            }
            # Create a new attendance record
            attendance_record = {
                'ID': iD,
                'Full_Name': " ".join(full_name),
                'Absent': 0
            }
            # Create a new exam record
            exam_record = {
                'ID': iD,
                'Full_Name': " ".join(full_name),
                'Monthly_Exam_1': None,
                'Mid_term_Exam': None,
                'Monthly_Exam_2': None,
                'Final_Exam': None,
                'Total': None
            }

            # Save the new records
            exam_data = pd.concat([exam_data, pd.DataFrame([exam_record])], ignore_index=True)
            with pd.ExcelWriter('Exam result.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                exam_data.to_excel(writer, sheet_name=sheet_name, index=False)

            attendance_data = pd.concat([attendance_data, pd.DataFrame([attendance_record])], ignore_index=True)
            with pd.ExcelWriter('Attendence.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                attendance_data.to_excel(writer, sheet_name=sheet_name, index=False)

            main_data = pd.concat([main_data, pd.DataFrame([new_student])], ignore_index=True)
            with pd.ExcelWriter('Registeration.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                main_data.to_excel(writer, index=False)

            i_s.append(iD)
            n.append(student_number)
            n2.append(guardian_number)

        while True:
            continue_prompt = input('Do you want to continue registration? [1. Yes / 2. No] ')
            if continue_prompt == '1':
                break
            elif continue_prompt == '2':
                continue_registration = False
                break
            else:
                print('Invalid input. Please try again.')

    print("Registration completed successfully.")
def Student_Data_Search():
    # Load data from Excel files
    registration_data = pd.read_excel('Registeration.xlsx', dtype={'Number_ka': str, 'Masuul_no': str})
    # Extract the necessary columns into lists
    student_ids = registration_data.ID.tolist()
    full_names = registration_data.Full_Name.tolist()
    phone_numbers = registration_data.Student_Number.tolist()
    genders = registration_data.Gender.tolist()
    guardian_names = registration_data.Guardian_Name.tolist()
    guardian_numbers = registration_data.Guardian_Number.tolist()
    birth_dates = registration_data.Date_of_Birth.tolist()
    registration_dates = registration_data.Registration_Date.tolist()
    classes = registration_data.Class.tolist()
    c_s = True
    while c_s:
        print('')
        student_id = input('Enter  ID Student: ') 
        if student_id in student_ids:
            index = student_ids.index(student_id)  # Find the index of the ID

            # Display student information
            print("\n--- Student Information ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'Student ID':<20}{student_id:<30}")
            print(f"{'Full_Name':<20}{full_names[index]:<30}")
            print(f"{'Student_Number':<20}{phone_numbers[index]:<30}")
            print(f"{'Gender':<20}{genders[index]:<30}")
            print(f"{'Guardian_Name':<20}{guardian_names[index]:<30}")
            print(f"{'Guardian_Number':<20}{guardian_numbers[index]:<30}")
            print(f"{'Date_of_Birth':<20}{birth_dates[index]:<30}")
            print(f"{'Registration_Date':<20}{registration_dates[index]:<30}")
            print(f"{'Class':<20}{classes[index]:<30}")
            print(f"{'-'*50}\n")
        else:
            print("This ID isn't registered. Try again.")
            Student_Data_Search()
        
        # Ask if the user wants to continue
        per = input("Do you want to continue student's search dataa [1.Yes | 2.No]: ")
        if per == '1':
            continue
        elif per == '2':
            c_s = False
        else:
            print('Enter 1 or 2')
def attendence():
    
    print('1.Form One\n2.Form Two\n3.Form Three\n4.Form Four')
    print('')

    while True:
        
        bb = input('Enter the class : ')
        if bb == '1':
            CLASS = 'Form1'
        elif bb == '2':
            CLASS = 'Form2'
        elif bb == '3':
            CLASS = 'Form3'
        elif bb == '4':
            CLASS = 'Form4'
        else:
            print('Unknown Class')
            continue
        xy = pd.read_excel('Attendence.xlsx', sheet_name=CLASS) 
        iD = xy.ID.tolist()
        name = xy.Full_Name.tolist()

        if not iD:
            print('This class is empty')
            continue
        attendance = []
        for i in range(len(name)):
            print('Enter attendance (P for Present, A for Absent):')
            while True:
                er = input(f'{iD[i]} {name[i]} : ')
                if er.capitalize() == 'P':
                    attendance.append(0) 
                    break
                elif er.capitalize() == 'A':
                    attendance.append(1)  
                    break
                else:
                    print('Soo gali kaliya A or P')
        xy['Absent'] = xy['Absent'] + attendance 
        with pd.ExcelWriter('Attendence.xlsx',mode='a',engine='openpyxl',if_sheet_exists='overlay')as pp:
            xy.to_excel(pp,sheet_name=CLASS,index=False)
        cont = input('Do you want to continue attendence? [1.Yes | 2.No]: ')
        if cont.capitalize() == '1':
            continue         
        elif cont.capitalize() == '2':
            break
        else:
            print("Enter 1 or 2")          
def Get_Exam_result():
    print('1.Form 1 \n2.Form 2 \n3.Form 3 \n4.Form 4')
    while True:
        Class=int(input("Enter the class: ".title()))
        if Class in [1,2,3,4]:
            try:
                if Class==1:
                    sheet_name = 'Form1'
                elif Class==2:
                    sheet_name='Form2'
                elif Class==3:
                    sheet_name='Form3'
                elif Class==4:
                    sheet_name='Form4'
                break
            except:
                    print("Enter Number Only")                
    exam_data = pd.read_excel('Exam result.xlsx',sheet_name=sheet_name)
    student_ids = exam_data.ID.tolist()
    full_names = exam_data.Full_Name.tolist()

    monthly_exam_1 = exam_data.Monthly_Exam_1.tolist()
    mid_term = exam_data.Mid_term_Exam.tolist()
    monthly_exam_2 = exam_data.Monthly_Exam_2.tolist()
    final_exam = exam_data.Final_Exam.tolist()


    while True:
        student_id = input('Enter the ID of the student: ').upper()
        if student_id in student_ids:
            index = student_ids.index(student_id)  # Find the index of the ID

            # Display student information
            print("--- Student Exam Result ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'Student ID':<20}{student_id:<30}")
            print(f"{'Name':<20}{full_names[index]:<30}")
            print(f"{'Monthly Exam 1':<20}{monthly_exam_1[index - 1]:<30}")
            print(f"{'Mid Term':<20}{mid_term[index -1]:<30}")
            print(f"{'Monthly Exam 2':<20}{monthly_exam_2[index-1]:<30}")
            print(f"{'Final Exam':<20}{final_exam[index-1]:<30}")
            # print(f"{'Total Marks':<20}{total_marks[index-1]:<30}")
            print(f"{'-'*50}\n")
            break
        else:
            print('This ID is not registered. Try again.')
        
        # Ask if the user wants to continue
    while True:
        repeat = input("Do you want to get another student's exam result (1. Yess | 2. No): ")
        if repeat == '1':
            Get_Exam_result()
        elif repeat == '2':
            exit()
        else:
            print('Enter 1 or 2')
def Teacher_registeration():
    # Load the data
    main_data = pd.read_excel('Teacher registeration.xlsx' , dtype={'Number': str})
    main_data_1=pd.read_excel('Registeration.xlsx' ,  dtype={'Student_Number': str})
    Student_Name = main_data_1.Full_Name.tolist()
    Student_Number= main_data_1.Student_Number.tolist()
    Teacher_ID = main_data.ID.tolist()
    Teacher_Number = main_data.Number.tolist() 
    Teacher_Name = main_data.Full_Name.tolist()
    degrees = [
    # Bachelor's Degrees
    "Bachelor Of Computer Science",
    "Bachelor Of Psychology",
    "Bachelor Of Business Administration",
    "Bachelor Of Civil Engineering",
    "Bachelor Of Medicine and Surgery",
    "Bachelor Of Medicine",
    "Bachelor Of Fine Arts",
    "Bachelor Of Environmental Science",
    "Bachelor Of Biotechnology",
    "Bachelor Of Sociology",
    "Bachelor Of Nursing",
    "Bachelor Of Mathematics",
    "Bachelor Of Physics",
    "Bachelor Of Chemistry",
    "Bachelor Of Political Science",
    "Bachelor Of Geology",
    "Bachelor Of History",
    "Bachelor Of English Literature",
    "Bachelor Of Zoology",
    "Bachelor Of Anthropology",
    "Bachelor Of Education",
    "Bachelor Of Agriculture",
    "Bachelor Of Journalism and Mass Communication",
    "Bachelor Of Statistics",
    "Bachelor Of Electronics",
    "Bachelor Of Philosophy",
    "Bachelor Of Visual Arts",
    "Bachelor Of Microbiology",
    "Bachelor Of Genetics",
    "Bachelor Of Marine Biology",
    "Bachelor Of Foreign Languages",
    "Bachelor Of Architecture",
    "Bachelor Of Fashion Design",
    "Bachelor Of Computer Engineering",
    "Bachelor Of Information Technology",
    "Bachelor Of Commerce",
    "Bachelor Of Laws",
    "Bachelor Of Forensic Science",
    "Bachelor Of Public Administration",
    "Bachelor Of Nutrition and Dietetics",
    "Bachelor Of Environmental Engineering",
    "Bachelor Of Robotics",
    "Bachelor Of Astronomy",
    "Bachelor Of Cybersecurity",
    "Bachelor Of Cloud Computing",
    "Bachelor Of Artificial Intelligence",
    "Bachelor Of Business Management",
    "Bachelor Of Tourism and Hospitality Management",
    "Bachelor Of Social Work",
    "Bachelor Of Data Analytics",
    "Bachelor Of Media Studies",
    "Bachelor Of Aerospace Engineering",
    "Bachelor Of Software Engineering",
    "Bachelor Of Theatre and Drama",
    "Bachelor Of Music",
    "Bachelor Of Renewable Energy",
    "Bachelor Of Supply Chain Management",
    "Bachelor Of Criminology",
    "Bachelor Of Sculpture",
    "Bachelor Of Digital Marketing",
    "Bachelor Of Health Informatics",
    "Bachelor Of Electronics and Communication Engineering",
    "Bachelor Of Mechanical Engineering",
    "Bachelor Of Biomedical Engineering",
    "Bachelor Of International Relations",
    "Bachelor Of Development Studies",
    "Bachelor Of Applied Physics",
    "Bachelor Of Cognitive Science",
    "Bachelor Of Oceanography",
    "Bachelor Of Wildlife Conservation",
    "Bachelor Of Film Studies",
    "Bachelor Of Game Development",
    "Bachelor Of Petroleum Engineering",
    "Bachelor Of Textile Engineering",
    "Bachelor Of Peace and Conflict Studies",
    "Bachelor Of Structural Engineering",
    "Bachelor Of Gender Studies",
    "Bachelor Of Civil Engineering",
    "Bachelor Of Classical Studies",
    "Bachelor Of Urban Studies",
    "Bachelor Of Nanotechnology",
    "Bachelor Of Music Therapy",
    "Bachelor Of Human Genetics",
    "Bachelor Of Environmental Health",
    "Bachelor Of Energy Engineering",
    "Bachelor Of Industrial Design",
    "Bachelor Of Metallurgical Engineering",
    "Bachelor Of Agricultural Engineering",
    "Bachelor Of Cultural Studies",
    "Bachelor Of Disaster Management",
    "Bachelor Of Game Design",
    "Bachelor Of Instrumentation Engineering",
    "Bachelor Of Sustainable Development",
    "Bachelor Of Adventure Tourism",
    "Bachelor Of Scriptwriting",
    "Bachelor Of Agronomy",
    "Bachelor Of Wildlife Biology",
    "Bachelor Of Computational Biology",
    "Bachelor Of Human Rights",
    "Bachelor Of Applied Mathematics",
    "Bachelor Of Visual Communication",
    
    # Master's Degrees
    "Master Of Artificial Intelligence",
    "Master Of International Relations",
    "Master Of Business Administration",
    "Master Of Public Health",
    "Master Of Robotics",
    "Master Of Creative Writing",
    "Master Of Cybersecurity",
    "Master Of Data Science",
    "Master Of Renewable Energy",
    "Master Of Journalism",
  
    # PhD Degrees
    "PhD Of Computer Science",
    "PhD Of Artificial Intelligence",
    "PhD Of Data Science",
    "PhD Of Environmental Science",
    "PhD Of Physics",
    "PhD Of Chemistry",
    "PhD Of Biology",
    "PhD Of Biotechnology",
    "PhD Of Mathematics",
    "PhD Of Mechanical Engineering",
    ]
    ma_r = True

    
    def generate_next_id(existing_ids):
            if not existing_ids:
                return 'MT001'
            max_id = max(int(i[2:]) for i in existing_ids if i.startswith('MT'))
            return f"MT{max_id + 1:03}" 
    # print(generate_next_id(Teacher_ID))



    while True:
        try:
            Number_of_Teachers=int(input("How many teachers you want to register: ".title()))
        except ValueError:
            print("Enter Number Only")
            continue

        for i in range(1,Number_of_Teachers+1):
            # Automatically generate the ID
            iD=generate_next_id(Teacher_ID)
            # Teacher_ID.append(iD)
            
            # Validate The Full Name
            while True:
                Name=input("Enter the full name of the teacher (First  Middle  Last): ".title()).title().strip()
                if Name in Teacher_Name or Name in Student_Name:  
                     print('This name is already registered'.title())
                elif all(len(i) >= 3 and i.isalpha() for i in Name.split()) and len(Name.split()) == 3:
                    # Teacher_Name.append(Name)
                    break
                else:
                    print("Enter the Full Name and Each Name Must contain at least 3 Letters")
                print("")

            # Validate Number
            while True:
                Number = input("Enter The Teacher's Number (061xxxxxxx): ")
                if Number in Teacher_Number or Number in Student_Number:
                    print('This number is already registered'.title())
                elif Number.startswith('061') and len(Number) == 10 and Number.isdigit():
                    # Teacher_Number.append(Number)
                    break
                else:
                    print("Wrong Number.")

            # Validate Sex
            while True:
                Sex = input("Enter Teacher's Sex (Male or Female): ").capitalize()
                if Sex in ['Male', 'Female']:
                    break
                else:
                    print("Enter Male or Female")

            # Validate Degree
            while True:
                Degree = input("Enter Teacher's Degree: ").title()
                if Degree in degrees:
                    break
                else:
                    print("Enter Valid Degree")
            
            # Validate Subject
            while True:
                Subject = input("Enter Teacher's Subject: ").title()
                if Subject in ['Math','Biology','Chemistry','Physics','Somali','English','ArabiC','Geography','History','Technoogy','Business','Islamic Education']:
                    print(f"{iD} {Name} is successfully registered")
                    print(f"The ID of Teacher {Name} is: {iD}")
                    break
                else:
                    print("Enter Valid Subject")

            # Saving to excel
            new_row = {
                'ID': iD,
                'Full_Name': Name,
                'Number': Number,
                'Sex': Sex,
                'Degree': Degree,
                'Subject':Subject,
                'Date': time.strftime('%d/%m/20%y')
            }
            main_data = pd.concat([main_data, pd.DataFrame([new_row])], ignore_index=True)

            # Save the updated data
            with pd.ExcelWriter('Teacher registeration.xlsx', mode='a', engine='openpyxl',if_sheet_exists = 'overlay') as ui:
                main_data.to_excel(ui, index=False)
            Teacher_ID.append(iD)
            Teacher_Name.append(Name)
            Teacher_Number.append(Number)
         
        break
def Teacher_Data_Search():
    # Load data from Excel files
    Teacher_regestration_data = pd.read_excel('Teacher registeration.xlsx', dtype={'Number': str})
    # exam_data = pd.read_excel('Exam result.xlsx')
    
    # Extract the necessary columns into lists
    Teacher_ids = Teacher_regestration_data.ID.tolist()
    full_names = Teacher_regestration_data.Full_Name.tolist()
    phone_numbers = Teacher_regestration_data.Number.tolist()
    Sexs = Teacher_regestration_data.Sex.tolist()
    Teachers_Degrees = Teacher_regestration_data.Degree.tolist()
    Subjects = Teacher_regestration_data.Subject.tolist()
    registration_dates = Teacher_regestration_data.Date.tolist()

    while True:
        Teacher_ID = input('Enter the ID of the Teacher: ').upper()
        if Teacher_ID in Teacher_ids:
            index = Teacher_ids.index(Teacher_ID)  # Find the index of the ID

            # Display student information
            print("\n--- Teacher Data ---")
            print('')
            print(f"{'Field':<20}{'Value':<30}")
            print(f"{'-'*50}")
            print(f"{'ID':<20}{Teacher_ID:<30}")
            print(f"{'Name':<20}{full_names[index]:<30}")
            print(f"{'Number':<20}{phone_numbers[index]:<30}")
            print(f"{'Sex':<20}{Sexs[index]:<30}")
            print(f"{'Degree':<20}{Teachers_Degrees[index]:<30}")
            print(f"{'Subject':<20}{Subjects[index]:<30}")
            print(f"{'Date':<20}{registration_dates[index]:<30}")
            print(f"{'-'*50}\n")
            break
        else:
            print('This ID is not registered. Try again.')
        
        # Ask if the user wants to continue
    while True:
        per = input("Do you want to get another teacher's data [1.yes/2.no]: ")
        if per == '1':
            continue
        elif per == '2':
            break
        else:
            print('Enter number only')
def Schedule():
    def F1():
        print(f" \
            \nSchedule Form one (F1)")
        print('''-------------------------------------------------------------------------------------------------------------------------------
    Maalin         | Period 1             | Period 2     | Period 3             | Period 4    | Period 5   | Period 6             | Period 7
    -------------------------------------------------------------------------------------------------------------------------------
    Sabti          | Islamic Education    | Biology      | English              | Somali      | History    | Arabic               | Chemistry  
    Axad           | History              | Business     | Islamic Education    | Physics     | English    | Geography            | Biology 
    Isniin         | Math                 | English      | Arabic               | Business    | Physics    | Geography            | Somali  
    Talaado        | Arabic               | Geography    | Technology           | Business    | Physics    | Islamic Education    | Chemistry  
    Arbaco         | Math                 | Geography    | History              | Somali      | English    | Arabic               | Technology
    -------------------------------------------------------------------------------------------------------------------------------''')
    def F2():
        print(f" \
            \nSchedule Form Two (F2)")
        print('''-----------------------------------------------------------------------------------------------------------------------------------
    Maalin         | Period 1             | Period 2             | Period 3      | Period       | Period 5     | Period 6             | Period 7
    -----------------------------------------------------------------------------------------------------------------------------------
    Sabti          | Chemistry            | History              | Arabic        | Somali       | Physics      | Islamic Education    | Biology
    Axad           | Geography            | Islamic Education    | Technology    | Business     | History      | Physics              | Chemistry
    Isniin         | Physics              | English              | Math          | Geography    | Chemistry    | Biology              | Technology
    Talaado        | English              | Arabic               | Technology    | Chemistry    | Somali       | Biology              | Physics
    Arbaco         | Islamic Education    | History              | Arabic        | English      | Biology      | Geography            | Somali
    -----------------------------------------------------------------------------------------------------------------------------------''')

    def F3():
        print(f" \
            \nSchedule Form Three (F3)")
        print('''-----------------------------------------------------------------------------------------------------------------
    Maalin         | Period 1             | Period 2      | Period 3     | Period 4      | Period 5     | Period 6    | Period 7
    -----------------------------------------------------------------------------------------------------------------
    Sabti          | History              | Geography     | Chemistry    | Technology    | Business     | Biology     | English
    Axad           | Biology              | Somali        | Business     | Arabic        | Physics      | Math        | English
    Isniin         | Islamic Education    | Technology    | Somali       | Biology       | Physics      | History     | English
    Talaado        | History              | Chemistry     | English      | Somali        | Geography    | Business    | Arabic
    Arbaco         | Geography            | History       | Physics      | Math          | English      | Arabic      | Chemistry
    -----------------------------------------------------------------------------------------------------------------''')
    def F4():
        print(f" \
            \nSchedule Form Four (F4)")
        print('''------------------------------------------------------------------------------------------------------------------------------
    Maalin         | Period 1   | Period 2  | Period 3  | Period 4  | Period 5  | Period 6  | Period 7
    ------------------------------------------------------------------------------------------------------------------------------
    Sabti          | Somali     | History              | Islamic Education    | Business             | Arabic      | Chemistry     | Physics
    Axad           | English    | Islamic Education    | Geography            | Somali               | Physics     | Arabic        | Chemistry
    Isniin         | Biology    | Math                 | English              | Chemistry            | History     | Technology    | Geography
    Talaado        | History    | Math                 | Technology           | Islamic Education    | Arabic      | English       | Somali
    Arbaco         | Arabic     | English              | Somali               | Technology           | Business    | Geography     | Math
    ------------------------------------------------------------------------------------------------------------------------------''')
    print("1.Form One (F1)\
        \n2.Form Two (F2)\
        \n3.Form Three (F3)\
        \n4.Form Four (F4) ")
    while True:
        try:
            Class=int(input("Enter the class you want: ".title()))
            if Class==1:
                F1()
            elif Class==2:
                F2()
            elif Class==3:
                F3()
            elif Class==4:
                F4()
            else:
                print("This class isn't available")
            break
        except:
            print("Enter Number Only")
def set_exam_result():
    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import load_workbook
    workbook = load_workbook("Exam result.xlsx")
    Monthly_Exam_1_Result = []
    Mid_term_Result = []
    Monthly_Exam_2_Result = []
    Final_Result = []
    total = []
    def main():
        global ID, Subject, Subjects, result, Class, Name, sheet_name
        Subjects = [i for i in range(1, 13)]
        ID = []
        result = []
        Name = []
        print('1.Math \n2.Biology \n3.Chemistry \n4.Physics \n5.Somali \n6.English \n7.Arabic \n8.Geography \n9.History \n10.Technology \n11. Business \n12.Islamic Education')
        while True:
            try:
                Subject = int(input("Enter the subject: ".title()))
                if Subject in Subjects:
                    print('1.Form 1 \n2.Form 2 \n3.Form 3 \n4.Form 4')
                    while True:
                        try:
                            Class = int(input("Enter the class: ".title()))
                            if Class in [1, 2, 3, 4]:
                                while True:
                                    try:
                                        Number_students = int(input("Enter the number of students you want to record the exam : ".title()))
                                        for i in range(1, Number_students + 1):
                                            sheet_name = f'Form{Class}'
                                            xy = pd.read_excel('Exam result.xlsx', sheet_name=sheet_name)
                                            S_ID = xy.ID.tolist()
                                            while True:
                                                # print(Student_ID)
                                                Student_ID = input(f"Enter the ID of student {i}: ")
                                                if Student_ID in S_ID:
                                                    if Student_ID not in ID:
                                                        ID.append(Student_ID)
                                                        Name.append(None)
                                                        Monthly_Exam_1_Result.append(None)
                                                        Mid_term_Result.append(None)
                                                        Monthly_Exam_2_Result.append(None)
                                                        Final_Result.append(None)
                                                        total.append(None)
                                                        break
                                                    else:
                                                        print(f"You Entered this ID {Student_ID}. Enter Another ID")
                                                else:
                                                    print("This ID is not registered. Try again.")
                                            while True:
                                                try:
                                                    student_result = int(input(f"Enter the result of student {i}: ".title()))
                                                    if student_result in Result:
                                                        result.append(student_result)
                                                        break
                                                    else:
                                                        print("Wrong Result")
                                                except:
                                                    print("Enter Number Only")
                                        break
                                    except:
                                        print("Enter Number Only")
                                break
                            else:
                                print("Unknown Class")
                        except:
                            print("Enter Number Only")
                    break
                else:
                    print("Unknown Subject")
            except:
                print("Enter Number Only")
    def Monthly_Exam_1():
        global Result, Monthly_Exam_1_Result
        Result = [i for i in range(21)]
        main()

    def Mid_term_Exam():
        global Result, Mid_term_Result
        Result = [i for i in range(31)]
        main()

    def Monthly_Exam_2():
        global Result, Monthly_Exam_2_Result
        Result = [i for i in range(21)]
        main()

    def Final_Exam():
        global Result, Final_Result
        Result = [i for i in range(31)]
        main()

    print("1.Monthly Exam 1\n2.Mid-Term Exam\n3.Monthly Exam 2\n4.Final Exam\n5.Exit")
    while True:
        try:
            Exam_Type = int(input("Enter the Exam_Type you want: ".title()))
            def exams():
                if Exam_Type == 1:
                    Monthly_Exam_1()
                elif Exam_Type == 2:
                    Mid_term_Exam()
                elif Exam_Type == 3:
                    Monthly_Exam_2()
                elif Exam_Type == 4:
                    Final_Exam()
                elif Exam_Type == 5:
                    print("Exiting exam result management system ".title())
                else:
                    print("Invalid choice".title())
            exams()
            break
        except:
            print("Enter Number Only")

    if Exam_Type == 1:
        Monthly_Exam_1_Result.clear()
        Monthly_Exam_1_Result.extend(result)
        E = 'Monthly_Exam_1'
    elif Exam_Type == 2:
        Mid_term_Result.clear()
        Mid_term_Result.extend(result)
        E = 'Mid_term_Exam'
    elif Exam_Type == 3:
        Monthly_Exam_2_Result.clear()
        Monthly_Exam_2_Result.extend(result)
        E = 'Monthly_Exam_2'
    elif Exam_Type == 4:
        Final_Result.clear()
        Final_Result.extend(result)
        E = 'Final_Exam'
    def Saving_excel():
        # Specify the file name
        file = 'Exam result.xlsx'

        # Load the data for the selected sheet
        df = pd.read_excel(file, sheet_name=sheet_name)

        # Prepare the new data to update or add
        data = {
            'ID': ID,
            'Monthly_Exam_1': Monthly_Exam_1_Result,
            'Mid_term_Exam': Mid_term_Result,
            'Monthly_Exam_2': Monthly_Exam_2_Result,
            'Final_Exam': Final_Result,
        }

        # Convert the data dictionary to a DataFrame
        df1 = pd.DataFrame(data)

        # Lists to track operations
        existing_students = []
        updated_students = []

        # Iterate through the new data
        for i, row in df1.iterrows():
            # Check if the student ID already exists in the sheet
            existing_row = df[df['ID'] == row['ID']]
            if not existing_row.empty:
                # If the specific exam result already exists
                if not pd.isna(existing_row.iloc[0][E]):
                    print(f"Result for {E} already exists for Student ID: {row['ID']}.")
                    print("")
                    user_input = input(f"Do you want to delete the existing result and enter a new one for Student ID: {row['ID_Number']}? (1. Yes | 1.No): ").strip()
                    if user_input == '1':
                        # Delete the existing result
                        df.loc[existing_row.index, E] = None
                        # Ask for a new result
                        while True:
                            try:
                                new_result = int(input(f"Enter the new result for Student ID: {row['ID']} in {E}: "))
                                # exams()
                                if new_result in Result:  # Replace with your valid range if different
                                    df.loc[existing_row.index, E] = new_result
                                    updated_students.append(row['ID'])
                                    break
                                else:
                                    print("Please enter a valid result.")
                            except ValueError:
                                print("Please enter a numeric value.")
                    else:
                        existing_students.append(row['ID'])
                else:
                    # If the result does not exist, update it
                    df.loc[existing_row.index, E] = row[E]
                    updated_students.append(row['ID'])
            else:
                # If the student ID does not exist
                print(f"Student ID: {row['ID']} does not exist in the sheet.")

        # Recalculate the 'Total' column by summing up all exam results for each student
        df['Total'] = (
            df[['Monthly_Exam_1', 'Mid_term_Exam', 'Monthly_Exam_2', 'Final_Exam']]
            .fillna(0)  # Replace NaN values with 0
            .sum(axis=1)  # Sum across columns
        )

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Print results
        if updated_students:
            print(f"Result(s) for Student IDs: {', '.join(updated_students)} successfully saved to {file} in the sheet '{sheet_name}'.")
        if existing_students:
            print(f"Result for {E} already exists for Student IDs: {', '.join(existing_students)}.")

    Saving_excel()
def del_s():
    main_data = pd.read_excel('Registeration.xlsx')
    att = pd.read_excel('Attendence.xlsx')
    exam = pd.read_excel('Exam result.xlsx')
    i_s = main_data.ID.tolist()
    name = main_data.Full_Name.tolist()
    while True:
        gf = True
        while gf:
                
            print('1.Form One\n2.Form Two\n3.Form Three\n4.Form Four')
            while True:
                try:
                    c = int(input('Enter a Class : '))
                    if c == 1:
                        sheetname = 'Form1'
                        break
                    elif c == 2:
                        sheetname = 'Form2'
                        break
                    elif c == 3:
                        sheetname = 'Form3'
                        break
                    elif c == 4:
                        sheetname = 'Form4'
                        break
                    else:
                        print('Please enter either 1 or 2 or 3 or 4') 
                except:
                    print('Only enter a number ') 
            while True:
                id_s = input('Enter Student ID : ') 
                if id_s in i_s and id_s in pd.read_excel('Attendence.xlsx',sheet_name=sheetname).ID.tolist() and pd.read_excel('Exam result.xlsx').ID.tolist():
                    while True:
                        a = int(input(f'Do you want to delete {name[i_s.index(id_s)]} [1.Yes/2.No]: '))
                        if a == 1:
                            break
                        elif a == 2:
                            break
                        else:
                            print('Please enter either 1 or 2')
                    gf = False
                    break
                else:
                    print('This ID is not registered')
                    gf = True 
                    break
        if a == 1:
            wb = load_workbook('Registeration.xlsx')
            ws = wb['Sheet1']

            # Read the Excel data into a DataFrame
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]  # Set the first row as column headers
            df = df[1:]  # Remove the header row from the data

            # Remove duplicates and the row with ID 'MS002'
            df = df.drop_duplicates()
            df = df[df['ID'] != id_s]

            # Clear the worksheet (except headers)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Write the cleaned DataFrame back to the worksheet
            for i, row in enumerate(df.values, start=2):
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)

            # Save the workbook
            wb.save('Registeration.xlsx')
            wb = load_workbook('Attendence.xlsx')
            ws = wb[sheetname]

            # Read the Excel data into a DataFrame
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]  # Set the first row as column headers
            df = df[1:]  # Remove the header row from the data

            # Remove duplicates and the row with ID 'MS002'
            df = df.drop_duplicates()
            df = df[df['ID'] != id_s]

            # Clear the worksheet (except headers)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Write the cleaned DataFrame back to the worksheet
            for i, row in enumerate(df.values, start=2):
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)

            # Save the workbook
            wb.save('Attendence.xlsx') 
            wb = load_workbook('Exam result.xlsx')
            ws = wb[sheetname]

            # Read the Excel data into a DataFrame
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]  # Set the first row as column headers
            df = df[1:]  # Remove the header row from the data

            # Remove duplicates and the row with ID 'MS002'
            df = df.drop_duplicates()
            df = df[df['ID'] != id_s]

            # Clear the worksheet (except headers)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Write the cleaned DataFrame back to the worksheet
            for i, row in enumerate(df.values, start=2):
                for j, value in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=value)

            # Save the workbook
            wb.save('Exam result.xlsx')

        else:
            pass
        cont = input('Do you want to continue [1.Yes | 2.No]: ')
        if cont.capitalize() == '1':
            continue         
        elif cont.capitalize() == '2':
            break
        else:
            print("Enter 1 or 2")
def login():
    print(f'\n{" " * 20}--- Al-anwaar Primary & Secondary School ---')
    print('')
    while True:
        username = input('Enter username: ').strip()
        password = input('Enter password : ').strip()
        print('\n')
        if username =='group_4' and password == 'Al_anwaar001':
            
            print('1.Student Management Sytem\n2.Teacher Management System\n3.Schedule Management System\n4.Exam Result Management System')
            print('')
            while True:
                ty = input('Enter one of them : ')
                if ty =='1':
                    print('1.Register Student \n2.Record attendance\n3.Get Student Data\n4.Delete Student') 
                    while True:
                        tt = input('Enter one of them : ')
                        if tt == '1':
                            register_students()
                            break
                        elif tt =='2':
                            attendence()
                            break
                        elif tt =='3':
                            Student_Data_Search()
                            break
                        elif tt =='4':
                            del_s()
                            break
                        else:
                            print('Invalid') 
                    break

                elif ty == '2':
                    print('1.Register Teacher\n2.Get Teacher Data')
                    while True:
                        tt = input('Enter one of them : ')
                        if tt =='1':
                            Teacher_registeration()
                            break
                        elif tt == '2':
                            Teacher_Data_Search()
                            break
                        else:
                            print('Invalid')
                    break
                elif ty == '3':
                    Schedule()
                    break
                elif ty == '4':
                    print('1.Record Exam result\n2.Get Exam result')
                    while True:
                        tt = input('Enter one of them : ')
                        if tt == '1':
                            set_exam_result()
                            break
                        elif tt == '2':
                            Get_Exam_result()
                            break
                        else:
                            print('Invalid')
                    break
                        
                else:
                    print('Invalid Choice') 
                break

        else:
            print('Username and password are invalid')        
login()  